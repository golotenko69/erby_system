from django.contrib.auth import get_user_model
from django.contrib.auth.middleware import get_user
from django.db.models import Q
from django.http import HttpResponseForbidden
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.shortcuts import render, get_object_or_404
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views import View
from datetime import timedelta
from django.http import HttpResponse
import openpyxl
import string
import random
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from erbu_main.forms import StudentForm, PassportForm, DisabilityForm, MseForm, PmpkForm, ParentFormSet, EducationInstitutionForm, EducationEndedForm, EducationProcessForm, EducationTargetForm, EmploymentForm
from erbu_main.mixins import AdminRequiredMixin, TeacherRequiredMixin
from erbu_main.models import Student, EducationTarget, DisabilityInfo, EducationProcess, EducationInstitution, StudentLog
from users.forms import CustomUserCreationForm
from django.http import JsonResponse

User = get_user_model()

def mark_logs_read(request):
    """Фоновая функция для сброса счетчика уведомлений"""
    if request.method == "POST" and request.user.is_authenticated and getattr(request.user, 'role', None) == 'ADMIN':
        # Находим все непрочитанные и делаем их прочитанными
        StudentLog.objects.filter(is_read=False).update(is_read=True)
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error"}, status=400)

# Вспомогательная функция для генерации строк
def generate_random_string(length=8):
    letters_and_digits = string.ascii_letters + string.digits
    return ''.join(random.choice(letters_and_digits) for i in range(length))

# Create your views here.
def index_view(request):
    return render(request, 'erbu_main/index.html')

def about_view(request):
    return render(request, 'erbu_main/about.html')

def students_view(request):
    return render(request, 'erbu_main/students.html')

def contacts_view(request):
    return render(request, 'erbu_main/contacts.html')

class StudentsListView(TeacherRequiredMixin, ListView):
    model = Student
    template_name = 'erbu_main/students/students.html'
    context_object_name = 'students'
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user

        if user.role == 'ADMIN':
            return Student.objects.all()

        if user.role == 'TEACHER':
            return Student.objects.filter(
                education_process__education_institution=user.education_institution
            ).distinct()

        return Student.objects.none()


class StudentCreateView(TeacherRequiredMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'erbu_main/students/students_form.html'
    success_url = reverse_lazy('students')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.request.POST:
            context['passport_form'] = PassportForm(
                self.request.POST or None
            )

            context['disability_form'] = DisabilityForm(
                self.request.POST or None
            )

            context['mse_form'] = MseForm(
                self.request.POST or None
            )

            context['pmpk_form'] = PmpkForm(
                self.request.POST or None
            )

            # formset только для родителей
            context['parent_formset'] = ParentFormSet(
                self.request.POST or None
            )

            # обычные формы
            context['edu_end_form'] = EducationEndedForm(
                self.request.POST or None
            )

            context['edu_proc_form'] = EducationProcessForm(
                self.request.POST or None,
                user=self.request.user
            )

            context['edu_tar_form'] = EducationTargetForm(
                self.request.POST or None
            )

            context['employment_form'] = EmploymentForm(
                self.request.POST or None
            )
        else:
            context['passport_form'] = PassportForm()
            context['disability_form'] = DisabilityForm()
            context['mse_form'] = MseForm()
            context['pmpk_form'] = PmpkForm()

            context['parent_formset'] = ParentFormSet()
            context['edu_end_form'] = EducationEndedForm()
            context['edu_proc_form'] = EducationProcessForm(user=self.request.user)
            context['edu_tar_form'] = EducationTargetForm()
            context['employment_form'] = EmploymentForm()

        return context

    def form_valid(self, form):
        context = self.get_context_data()

        passport_form = context['passport_form']
        disability_form = context['disability_form']
        mse_form = context['mse_form']
        pmpk_form = context['pmpk_form']

        parent_formset = context['parent_formset']
        edu_end_form = context['edu_end_form']
        edu_proc_form = context['edu_proc_form']
        edu_tar_form = context['edu_tar_form']
        employment_form = context['employment_form']

        # Проверяем валидность всех форм
        forms_valid = (
                form.is_valid() and
                passport_form.is_valid() and
                disability_form.is_valid() and
                mse_form.is_valid() and
                pmpk_form.is_valid() and
                parent_formset.is_valid() and
                edu_end_form.is_valid() and
                edu_proc_form.is_valid() and
                edu_tar_form.is_valid() and
                employment_form.is_valid()
        )

        if forms_valid:
            # Сохраняем студента
            student = form.save()

            # Сохраняем OneToOne поля
            passport = passport_form.save(commit=False)
            passport.student = student
            passport.save()

            disability = disability_form.save(commit=False)
            disability.student = student
            disability.save()

            mse = mse_form.save(commit=False)
            mse.disability = disability
            mse.save()

            pmpk = pmpk_form.save(commit=False)
            pmpk.disability = disability
            pmpk.save()

            education_ended = edu_end_form.save(commit=False)
            education_ended.student = student
            education_ended.save()

            education_process = edu_proc_form.save(commit=False)
            education_process.student = student

            if self.request.user.role == 'TEACHER':
                education_process.education_institution = (
                    self.request.user.education_institution
                )

            education_process.save()

            education_target = edu_tar_form.save(commit=False)
            education_target.student = student
            education_target.save()

            employment = employment_form.save(commit=False)
            employment.student = student
            employment.save()

            # Сохраняем formsets
            parent_formset.instance = student
            parent_formset.save()

            # ЛОГИРОВАНИЕ ДОБАВЛЕНИЯ
            user_institution = getattr(self.request.user, 'education_institution', None)
            StudentLog.objects.create(
                action='CREATE',
                student_name=f"{student.last_name} {student.first_name} {student.middle_name or ''}".strip(),
                user=self.request.user,
                institution_name=user_institution.name if user_institution else "Администрация"
            )

            messages.success(
                self.request,
                f"Студент '{student.first_name} {student.last_name}' успешно создан!"
            )
            return redirect('student_detail', student_id=student.pk)

        return self.render_to_response(self.get_context_data(form=form))

class StudentUpdateView(TeacherRequiredMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'erby/students/student_form.html'
    success_url = reverse_lazy('students')

    def form_valid(self, form):
        student = form.save()
        # ЛОГИРОВАНИЕ ОБНОВЛЕНИЯ
        user_institution = getattr(self.request.user, 'education_institution', None)
        StudentLog.objects.create(
            action='UPDATE',
            student_name=f"{student.last_name} {student.first_name} {student.middle_name or ''}".strip(),
            user=self.request.user,
            institution_name=user_institution.name if user_institution else "Администрация"
        )
        messages.success(self.request, f"Студент '{form.instance.first_name} {form.instance.last_name}' успешно обновлен!")
        return super().form_valid(form)


class StudentDeleteView(TeacherRequiredMixin, DeleteView):
    model = Student
    template_name = 'erbu_main/students/student_confirm_delete.html'
    success_url = reverse_lazy('students')
    pk_url_kwarg = 'student_id'

    def post(self, request, *args, **kwargs):
        obj = self.get_object()

        # ЛОГИРОВАНИЕ УДАЛЕНИЯ
        user_institution = getattr(self.request.user, 'education_institution', None)
        StudentLog.objects.create(
            action='DELETE',
            student_name=f"{obj.last_name} {obj.first_name} {obj.middle_name or ''}".strip(),
            user=self.request.user,
            institution_name=user_institution.name if user_institution else "Администрация"
        )

        messages.success(self.request, f"Студент '{obj.first_name} {obj.last_name}' успешно удален!")
        return super().delete(request, *args, **kwargs)

class EducationDeleteView(AdminRequiredMixin, DeleteView):
    model = EducationTarget
    template_name = 'erby/teachers/teacher_confirm_delete.html'
    success_url = reverse_lazy('teachers')
    pk_url_kwarg = 'education_id'

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        messages.success(self.request, f"Учитель {obj.user.first_name} {obj.user.last_name} успешно удален!")
        return super().delete(request, *args, **kwargs)\

class StudentDetailView(TeacherRequiredMixin, DetailView):
    model = Student
    template_name = 'erbu_main/students/student_detail.html'
    context_object_name = 'student'
    pk_url_kwarg = 'student_id'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object

        # Паспорт
        context['passport'] = getattr(student, 'passport', None)

        # Инвалидность и связанные данные
        try:
            disability = student.disability_info
            context['disability'] = disability
            context['mse'] = getattr(disability, 'mse', None)
            context['pmpk'] = getattr(disability, 'pmpk', None)
        except DisabilityInfo.DoesNotExist:
            context['disability'] = None
            context['mse'] = None
            context['pmpk'] = None

        # Образование
        context['education_ended'] = getattr(student, 'education_ended', None)
        context['education_process'] = getattr(student, 'education_process', None)
        context['education_target'] = getattr(student, 'education_target', None)
        # Родители
        context['parents'] = student.parents.all()

        # Трудоустройство
        context['employment'] = getattr(student, 'employment', None)

        return context

    def get_queryset(self):
        user = self.request.user

        if user.role == 'ADMIN':
            return Student.objects.all()

        return Student.objects.filter(
            Q(education_process__education_institution=user.education_institution) |
            Q(pk=self.kwargs.get('student_id'))
        ).distinct()

class SignUpView(CreateView):
    form_class = CustomUserCreationForm
    success_url = reverse_lazy('login')
    template_name = 'registration/signup.html'


class EducationInstitutionListView(AdminRequiredMixin, ListView):
    """Список всех учебных заведений"""
    model = EducationInstitution
    template_name = 'erbu_main/educations/institution_list.html'
    context_object_name = 'institutions'
    paginate_by = 20

    def get_queryset(self):
        return EducationInstitution.objects.all().order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Передаем роль в контекст, чтобы делать проверки в шаблоне
        context['is_admin'] = getattr(user, 'role', None) == 'ADMIN'

        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

        for institution in context['institutions']:
            institution.teacher = User.objects.filter(
                education_institution=institution,
                role='TEACHER'
            ).first()

        for institution in context['institutions']:
            # Базовый запрос процессов обучения для этого учреждения
            edu_processes = EducationProcess.objects.filter(education_institution=institution)

            # 1. Общее количество уникальных студентов
            institution.student_count = edu_processes.values('student').distinct().count()

            # Данные, доступные только для администратора
            if context['is_admin']:
                # 2. Сколько студентов добавлено за сегодня
                institution.students_added_today = edu_processes.filter(
                    created_at__gte=today_start
                ).values('student').distinct().count()

                # 3. Последние добавленные студенты (например, последние 3) для вывода времени
                institution.recent_additions = edu_processes.select_related('student').order_by('-created_at')[:3]

        return context

class EducationInstitutionDetailView(AdminRequiredMixin, DetailView):
    """Детальная информация об учебном заведении"""
    model = EducationInstitution
    template_name = 'erbu_main/educations/institution_detail.html'
    context_object_name = 'institution'
    pk_url_kwarg = 'institution_id'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        institution = self.object

        context['responsible_teacher'] = User.objects.filter(
            education_institution=institution,
            role='TEACHER'
        ).first()

        # Получаем всех студентов, обучающихся в этом учреждении
        education_processes = EducationProcess.objects.filter(
            education_institution=institution
        ).select_related('student').order_by('student__last_name', 'student__first_name')

        # Группируем по курсам
        students_by_course = {
            1: [], 2: [], 3: [], 4: []
        }

        for edu in education_processes:
            if edu.course in students_by_course:
                students_by_course[edu.course].append({
                    'student': edu.student,
                    'profession': edu.profession,
                    'form': edu.get_form_display(),
                    'grad_date': edu.grad_date
                })

        context['students_by_course'] = students_by_course
        context['total_students'] = education_processes.values('student').distinct().count()
        context['professions'] = education_processes.values_list('profession', flat=True).distinct()

        return context


class EducationInstitutionCreateView(AdminRequiredMixin, CreateView):
    """Создание нового учебного заведения и пользователя-учителя"""
    model = EducationInstitution
    form_class = EducationInstitutionForm
    template_name = 'erbu_main/educations/institution_form.html'
    success_url = reverse_lazy('institutions')

    def form_valid(self, form):
        # Сохраняем само учебное заведение
        response = super().form_valid(form)
        institution = self.object

        # Генерируем уникальный логин и пароль
        username = f"user_{generate_random_string(6).lower()}"
        password = generate_random_string(10)

        # Создаем связанного пользователя (учителя)
        User.objects.create_user(
            username=username,
            password=password,
            first_name=form.cleaned_data['teacher_first_name'],
            last_name=form.cleaned_data['teacher_last_name'],
            middle_name=form.cleaned_data.get('teacher_middle_name', ''),
            email=form.cleaned_data['teacher_email'],
            phone=form.cleaned_data['teacher_phone'],
            role='TEACHER',
            education_institution=institution,
            raw_password=password  # Сохраняем пароль, чтобы показать его админу
        )

        messages.success(
            self.request,
            f"Учреждение '{institution.name}' и ответственный успешно созданы!"
        )
        return response


class EducationInstitutionUpdateView(AdminRequiredMixin, UpdateView):
    """Редактирование учебного заведения"""
    model = EducationInstitution
    form_class = EducationInstitutionForm
    template_name = 'erbu_main/educations/institution_form.html'
    pk_url_kwarg = 'institution_id'

    def get_success_url(self):
        return reverse_lazy('institution_detail', kwargs={'institution_id': self.object.pk})

    def form_valid(self, form):
        messages.success(
            self.request,
            f"Учебное заведение '{form.instance.name}' успешно обновлено!"
        )
        return super().form_valid(form)


class EducationInstitutionDeleteView(AdminRequiredMixin, DeleteView):
    """Удаление учебного заведения с ограничением по студентам"""
    model = EducationInstitution
    template_name = 'erbu_main/educations/institution_confirm_delete.html'
    pk_url_kwarg = 'institution_id'
    success_url = reverse_lazy('institutions')
    context_object_name = 'institution'

    def dispatch(self, request, *args, **kwargs):
        """Проверяем наличие студентов до выполнения любого (GET/POST) запроса"""
        institution = self.get_object()
        students_count = EducationProcess.objects.filter(
            education_institution=institution
        ).values('student').distinct().count()

        # Если это POST-запрос (попытка удалить) и студенты есть — блокируем
        if request.method == 'POST' and students_count > 0:
            messages.error(
                request,
                f"Ошибка! Невозможно удалить учреждение '{institution.name}', так как в нём обучается {students_count} студентов."
            )
            return redirect('institution_detail', institution_id=institution.pk)

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        """Передаем количество студентов в контекст шаблона для динамического UI"""
        context = super().get_context_data(**kwargs)
        context['students_count'] = EducationProcess.objects.filter(
            education_institution=self.object
        ).values('student').distinct().count()
        return context

    def delete(self, request, *args, **kwargs):
        """Вызывается, если проверка в dispatch пройдена успешно и студентов нет"""
        institution = self.get_object()

        # Находим учителей этого заведения и деактивируем их, чтобы они не могли войти
        teachers = User.objects.filter(education_institution=institution, role='TEACHER')
        for teacher in teachers:
            teacher.is_active = False
            teacher.education_institution = None  # Стираем связь с удаленным учреждением
            teacher.save()

        messages.success(
            request,
            f"Учебное заведение '{institution.name}' успешно удалено. Связанные аккаунты сотрудников заблокированы."
        )
        return super().delete(request, *args, **kwargs)


class StudentFullEditView(UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'erbu_main/students/students_form.html'
    pk_url_kwarg = 'student_id'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)

        # Проверка доступа для учителя
        if self.request.user.role == 'TEACHER':
            # ИСПРАВЛЕНО: Проверяем через связанный объект education_process
            student_institution = None
            if hasattr(obj, 'education_process') and obj.education_process:
                student_institution = obj.education_process.education_institution

            if student_institution != self.request.user.education_institution:
                raise HttpResponseForbidden("Нет доступа")

        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object if self.object.pk else None

        if self.request.POST:
            context['passport_form'] = PassportForm(
                self.request.POST,
                instance=getattr(student, 'passport', None)
            )
            context['disability_form'] = DisabilityForm(
                self.request.POST,
                instance=getattr(student, 'disability_info', None)
            )

            # Для Mse и Pmpk нужно проверить существование disability
            disability = getattr(student, 'disability_info', None)
            mse_instance = disability.mse if disability and hasattr(disability, 'mse') else None
            pmpk_instance = disability.pmpk if disability and hasattr(disability, 'pmpk') else None

            context['mse_form'] = MseForm(self.request.POST, instance=mse_instance)
            context['pmpk_form'] = PmpkForm(self.request.POST, instance=pmpk_instance)

            context['parent_formset'] = ParentFormSet(self.request.POST, instance=student)
            context['edu_end_form'] = EducationEndedForm(
                self.request.POST,
                instance=getattr(student, 'education_ended', None)
            )

            context['edu_proc_form'] = EducationProcessForm(
                self.request.POST,
                instance=getattr(student, 'education_process', None),
                user=self.request.user
            )

            context['edu_tar_form'] = EducationTargetForm(
                self.request.POST,
                instance=getattr(student, 'education_target', None)
            )

            context['employment_form'] = EmploymentForm(
                self.request.POST,
                instance=getattr(student, 'employment', None)
            )
        else:
            # GET запрос - загружаем существующие данные
            disability = getattr(student, 'disability_info', None)
            mse_instance = disability.mse if disability and hasattr(disability, 'mse') else None
            pmpk_instance = disability.pmpk if disability and hasattr(disability, 'pmpk') else None

            context['passport_form'] = PassportForm(instance=getattr(student, 'passport', None))
            context['disability_form'] = DisabilityForm(instance=disability)
            context['mse_form'] = MseForm(instance=mse_instance)
            context['pmpk_form'] = PmpkForm(instance=pmpk_instance)

            context['parent_formset'] = ParentFormSet(instance=student)

            # --- ИСПРАВЛЕНО ТУТ: Передаем правильные связанные инстансы вместо student ---
            context['edu_end_form'] = EducationEndedForm(
                instance=getattr(student, 'education_ended', None)
            )
            context['edu_proc_form'] = EducationProcessForm(
                instance=getattr(student, 'education_process', None),
                user=self.request.user
            )
            context['edu_tar_form'] = EducationTargetForm(
                instance=getattr(student, 'education_target', None)
            )
            context['employment_form'] = EmploymentForm(
                instance=getattr(student, 'employment', None)
            )

        return context

    def form_valid(self, form):
        context = self.get_context_data()

        passport_form = context['passport_form']
        disability_form = context['disability_form']
        mse_form = context['mse_form']
        pmpk_form = context['pmpk_form']
        parent_formset = context['parent_formset']
        edu_end_form = context['edu_end_form']
        edu_proc_form = context['edu_proc_form']
        edu_tar_form = context['edu_tar_form']
        employment_form = context['employment_form']

        forms_valid = (
                form.is_valid() and passport_form.is_valid() and
                disability_form.is_valid() and mse_form.is_valid() and
                pmpk_form.is_valid() and parent_formset.is_valid() and
                edu_end_form.is_valid() and edu_proc_form.is_valid() and
                edu_tar_form.is_valid() and employment_form.is_valid()
        )

        if forms_valid:
            # Получаем текущее состояние студента из базы ДО сохранения формы
            old_student = Student.objects.get(pk=self.object.pk)
            changes_list = []

            # Словарь человекочитаемых названий полей
            field_labels = {
                'status': 'Статус', 'phone': 'Телефон', 'email': 'Email',
                'education_type': 'Тип оконч. образования', 'education_document': 'Документ оконч. образования',
                'agreement': 'Целевой договор', 'abilimpiks': 'Участие в Абилимпикс',
                'employment_status': 'Трудоустроен', 'accounting_employment': 'Состоит в ЦЗН',
                'resume_status': 'Резюме создано'
            }

            # Функция для сравнения полей формы и старой модели
            def check_changes(form_obj, old_obj, fields):
                for field in fields:
                    if field in form_obj.cleaned_data:
                        new_val = form_obj.cleaned_data[field]
                        old_val = getattr(old_obj, field, None)

                        if hasattr(form_obj.fields[field], 'choices') and new_val:
                            new_val = dict(form_obj.fields[field].choices).get(new_val, new_val)
                        if old_obj and hasattr(old_obj, f'get_{field}_display'):
                            old_val = getattr(old_obj, f'get_{field}_display')()

                        if str(old_val) != str(new_val) and (old_val or new_val):
                            label = field_labels.get(field, field)
                            changes_list.append(f"• {label}: «{old_val or 'нет'}» ➔ «{new_val or 'нет'}»")

            # Проверяем изменения по разным формам
            check_changes(form, old_student, ['status', 'phone', 'email'])

            if hasattr(old_student, 'education_ended'):
                check_changes(edu_end_form, old_student.education_ended, ['education_type', 'education_document'])
            if hasattr(old_student, 'education_target'):
                check_changes(edu_tar_form, old_student.education_target, ['agreement', 'abilimpiks'])
            if hasattr(old_student, 'employment'):
                check_changes(employment_form, old_student.employment,
                              ['employment_status', 'accounting_employment', 'resume_status'])

            # Сохраняем данные
            student = form.save()
            passport_form.save()
            disability_form.save()
            mse_form.save()
            pmpk_form.save()
            edu_end_form.save()
            edu_proc_form.save()
            edu_tar_form.save()
            employment_form.save()
            parent_formset.save()

            # Собираем итоговую строку изменений
            changes_text = "\n".join(changes_list) if changes_list else "Существенных изменений в ключевых полях нет."

            # ЛОГИРОВАНИЕ С ДЕТАЛИЗАЦИЕЙ
            user_institution = getattr(self.request.user, 'education_institution', None)
            StudentLog.objects.create(
                action='UPDATE',
                student_name=f"{student.last_name} {student.first_name} {student.middle_name or ''}".strip(),
                user=self.request.user,
                institution_name=user_institution.name if user_institution else "Администрация",
                changes=changes_text
            )

            messages.success(self.request, "Студент успешно сохранен")
            return redirect('student_detail', student_id=student.pk)

        return self.render_to_response(self.get_context_data(form=form))


class ExportStudentsExcelView(View):
    def post(self, request, *args, **kwargs):
        # 1. Получаем выбранные разделы из формы (чекбоксы)
        export_sections = request.POST.getlist('sections')

        # Если ничего не выбрано, по умолчанию берем базовую информацию
        if not export_sections:
            export_sections = ['base']

        # 2. Фильтруем студентов в зависимости от роли пользователя (как в вашей StudentsListView)
        user = request.user
        if user.role == 'ADMIN':
            queryset = Student.objects.all()
        elif user.role == 'TEACHER':
            queryset = Student.objects.filter(
                education_process__education_institution=user.education_institution
            ).distinct()
        else:
            queryset = Student.objects.none()

        # Оптимизируем запросы с помощью select_related / prefetch_related
        queryset = queryset.select_related(
            'passport', 'disability_info', 'disability_info__mse',
            'disability_info__pmpk', 'education_ended',
            'education_process', 'education_process__education_institution',
            'education_target', 'employment'
        ).prefetch_related('parents')

        # 3. Инициализируем Excel книгу
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Студенты"

        # Стилизация
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 4. Динамически формируем шапку таблицы (Headers)
        headers = []

        # Базовая информация (всегда присутствует или по выбору)
        if 'base' in export_sections:
            headers.extend(
                ['ID', 'Фамилия', 'Имя', 'Отчество', 'Статус', 'Пол', 'Email', 'Телефон', 'Дата рождения', 'СНИЛС',
                 'ИНН'])

        if 'passport' in export_sections:
            headers.extend(['Паспорт Серия', 'Паспорт Номер', 'Кем выдан', 'Дата выдачи', 'Код подр.', 'Место рождения',
                            'Место жительства'])

        if 'disability' in export_sections:
            headers.extend(['Статус ОВЗ', 'Группа инвалидности', 'Нозология', 'Дата снятия', 'МСЭ Серия', 'МСЭ Номер',
                            'МСЭ Дата выдачи', 'МСЭ Переосвидетельствование', 'ПМПК Номер', 'ПМПК Дата',
                            'Реком. программа'])

        if 'education' in export_sections:
            headers.extend(['Предыд. образование', 'Учреждение (предыд.)', 'Документ об образ.', 'Серия/Номер док.',
                            'Дата выдачи док.', 'Текущее учреждение', 'Профессия', 'Курс', 'Форма обучения',
                            'Срок обучения', 'Дата выпуска'])

        if 'employment' in export_sections:
            headers.extend(
                ['Трудоустроен', 'Место работы', 'Должность', 'Дата приема', 'Причина нетрудоустр.', 'Состоит в ЦЗН',
                 'Резюме создано'])

        if 'parents' in export_sections:
            headers.extend(['Родители (ФИО, Телефон, Email)'])

        ws.append(headers)

        # Применяем стили к шапке
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        # 5. Заполняем данными
        for row_num, student in enumerate(queryset, start=2):
            row_data = []

            if 'base' in export_sections:
                row_data.extend([
                    student.student_id, student.last_name, student.first_name, student.middle_name,
                    student.get_status_display(), student.get_sex_display(), student.email, student.phone,
                    student.birthday.strftime('%d.%m.%Y') if student.birthday else '', student.snils, student.inn
                ])

            if 'passport' in export_sections:
                p = getattr(student, 'passport', None)
                row_data.extend([
                    p.series if p else '', p.number if p else '', p.issued if p else '',
                    p.date_issued.strftime('%d.%m.%Y') if p and p.date_issued else '',
                    p.department_code if p else '', p.place_birth if p else '', p.place_resident if p else ''
                ])

            if 'disability' in export_sections:
                d = getattr(student, 'disability_info', None)
                mse = getattr(d, 'mse', None) if d else None
                pmpk = getattr(d, 'pmpk', None) if d else None
                row_data.extend([
                    d.get_status_ovz_display() if d else '', d.disability_group if d else '',
                    d.nosology_type if d else '',
                    d.year_removal.strftime('%d.%m.%Y') if d and d.year_removal else '',
                    mse.series if mse else '', mse.number if mse else '',
                    mse.date_issued.strftime('%d.%m.%Y') if mse and mse.date_issued else '',
                    mse.date_next_examination.strftime('%d.%m.%Y') if mse and mse.date_next_examination else '',
                    pmpk.number if pmpk else '',
                    pmpk.date_issued.strftime('%d.%m.%Y') if pmpk and pmpk.date_issued else '',
                    pmpk.education_programm_pmpk if pmpk else ''
                ])

            if 'education' in export_sections:
                ee = getattr(student, 'education_ended', None)
                ep = getattr(student, 'education_process', None)
                row_data.extend([
                    ee.get_education_type_display() if ee else '', ee.name if ee else '',
                    ee.get_education_document_display() if ee else '',
                    ee.series if ee else '', ee.date_issued.strftime('%d.%m.%Y') if ee and ee.date_issued else '',
                    ep.education_institution.name if ep and ep.education_institution else '',
                    ep.profession if ep else '',
                    ep.course if ep else '', ep.get_form_display() if ep else '', ep.term if ep else '',
                    ep.grad_date.strftime('%d.%m.%Y') if ep and ep.grad_date else ''
                ])

            if 'employment' in export_sections:
                em = getattr(student, 'employment', None)
                row_data.extend([
                    em.get_employment_status_display() if em else '', em.place_job if em else '',
                    em.position if em else '',
                    em.hiring_date.strftime('%d.%m.%Y') if em and em.hiring_date else '',
                    em.reason_not_employment if em else '',
                    em.get_accounting_employment_display() if em else '', em.get_resume_status_display() if em else ''
                ])

            if 'parents' in export_sections:
                parents_list = []
                for parent in student.parents.all():
                    p_info = f"{parent.last_name or ''} {parent.first_name or ''} {parent.middle_name or ''}".strip()
                    p_contacts = f"Тел: {parent.phone or '-'}, Email: {parent.email or '-'}"
                    parents_list.append(f"{p_info} ({p_contacts})")
                row_data.append("\n".join(parents_list) if parents_list else "Нет данных")

            ws.append(row_data)

            # Стилизация строк данных
            ws.row_dimensions[row_num].height = 22
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = align_left if cell.column != 1 else align_center

        # Авто-выравнивание ширины колонок под контент
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # 6. Подготовка HTTP-ответа для скачивания файла
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="students_report.xlsx"'
        wb.save(response)
        return response


class LogDetailJsonView(View):
    def get(self, request, log_id, *args, **kwargs):
        if not request.user.is_authenticated or getattr(request.user, 'role', None) != 'ADMIN':
            return JsonResponse({'error': 'Доступ запрещен'}, status=403)

        log_item = get_object_or_404(StudentLog, pk=log_id)

        data = {
            'action_display': log_item.get_action_display(),
            'student_name': log_item.student_name,
            'user': f"{log_item.user.last_name or ''} {log_item.user.first_name or ''}".strip() if log_item.user else "Система",
            'institution': log_item.institution_name or "Не указано",
            'date': log_item.created_at.strftime('%d.%m.%Y в %H:%M'),
            'changes': log_item.changes or "Нет подробных данных"
        }
        return JsonResponse(data)